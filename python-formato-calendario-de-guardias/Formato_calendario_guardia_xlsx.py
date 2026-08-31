#!/usr/bin/env python3
"""
Generador del CALENDARIO DE GUARDIA NOCTURNA — versión hoja de cálculo (.xlsx)
Primera Compañía de Bomberos - Bomba Chile, Coquimbo

Editable directamente en LibreOffice Calc / Excel: cambia los datos en el
diccionario SEMANAS más abajo (o directamente en las celdas del archivo
generado) y las celdas de color/combinadas ya quedan listas.

Uso:
    python3 calendario_guardia_xlsx.py
Salida: calendario_guardia.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — edita esto cada mes/periodo
# ============================================================
TITULO = "Guardia Nocturna (08 de Junio al 05 de julio del 2026)"

SEMANAS = [
    {
        "color": "roja",
        "dias": [
            {"nombre": "LUNES", "fecha": "08-06-26", "obac": "41", "maquinista": "Cristian Silva",
             "voluntarios": ["Aymara Segura", "Karla Inzunza", "Ian Duque", "C. Gamboa"]},
            {"nombre": "MARTES", "fecha": "09-06-26", "obac": "41", "maquinista": "Cristian Silva",
             "voluntarios": ["Mikaela Silva", "Arlett Adaros", "C. Adaros"]},
            {"nombre": "MIÉRCOLES", "fecha": "10-06-26", "obac": "103", "maquinista": "Matias Nuñez",
             "voluntarios": ["Cristian Silva", "E. Jopia", "C. Gamboa"]},
            {"nombre": "JUEVES", "fecha": "11-06-26", "obac": "101", "maquinista": "R. Rodriguez",
             "voluntarios": ["Ian Duque", "P. Aracena", "C. Adaros", "Arlett Adaros"]},
            {"nombre": "VIERNES", "fecha": "12-06-26", "obac": "101", "maquinista": "A. Santana",
             "voluntarios": ["Matias Nuñez", "C. Ramirez", "E. Jopia", "P. Aracena", "P. Saavedra", "Felipe Melo"]},
            {"nombre": "SÁBADO", "fecha": "13-06-26", "obac": "Aliro Albanez", "maquinista": "Marcelo Muñoz",
             "voluntarios": ["C. Ramirez", "E. Jopia", "C. Gamboa"]},
            {"nombre": "DOMINGO", "fecha": "14-06-26", "obac": "Aliro Albanez", "maquinista": "Cristian Silva",
             "voluntarios": ["Emmanuel Jopia", "C. Gamboa"]},
        ],
        "oficial_de": "102",
    },
    {
        "color": "azul",
        "dias": [
            {"nombre": "LUNES", "fecha": "15-06-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "MARTES", "fecha": "16-06-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "MIÉRCOLES", "fecha": "16-06-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "JUEVES", "fecha": "17-06-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "VIERNES", "fecha": "18-06-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "SÁBADO", "fecha": "19-06-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "DOMINGO", "fecha": "20-06-26", "obac": "", "maquinista": "", "voluntarios": []},
        ],
        "oficial_de": "101",
    },
    {
        "color": "roja",
        "dias": [
            {"nombre": "LUNES", "fecha": "22-06-26", "obac": "103", "maquinista": "Matias Nuñez",
             "voluntarios": ["Aymara Segura", "Franco Rangel"]},
            {"nombre": "MARTES", "fecha": "23-06-26", "obac": "Aliro Albanez", "maquinista": "102",
             "voluntarios": ["Karla Inzunza", "Mikaela Silva", "Ian Duque", "Patricio Varela"]},
            {"nombre": "MIÉRCOLES", "fecha": "24-01-26", "obac": "103", "maquinista": "Matias Nuñez",
             "voluntarios": ["Ian Duque", "Franco Rangel"]},
            {"nombre": "JUEVES", "fecha": "25-06-26", "obac": "103", "maquinista": "A. Santana",
             "voluntarios": ["P. Varela", "Franco Rangel", "C. Adaros", "Arlett Adaros"]},
            {"nombre": "VIERNES", "fecha": "26-06-26", "obac": "C. Ramirez", "maquinista": "102",
             "voluntarios": ["Arlett Adaros", "C. Adaros", "P. Saavedra", "Felipe Melo"]},
            {"nombre": "SÁBADO", "fecha": "27-06-26", "obac": "Aliro Albanez", "maquinista": "Marcelo Muñoz",
             "voluntarios": ["Mikaela Silva", "P. Varela", "P. Saavedra", "Felipe Melo"]},
            {"nombre": "DOMINGO", "fecha": "28-06-26", "obac": "C. Ramirez", "maquinista": "41",
             "voluntarios": ["P. Varela", "P. Saavedra"]},
        ],
        "oficial_de": "103",
    },
    {
        "color": "azul",
        "dias": [
            {"nombre": "LUNES", "fecha": "29-06-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "MARTES", "fecha": "30-06-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "MIÉRCOLES", "fecha": "01-07-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "JUEVES", "fecha": "02-07-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "VIERNES", "fecha": "03-07-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "SÁBADO", "fecha": "04-07-26", "obac": "", "maquinista": "", "voluntarios": []},
            {"nombre": "DOMINGO", "fecha": "05-07-26", "obac": "", "maquinista": "", "voluntarios": []},
        ],
        "oficial_de": "41",
    },
]

ROJO_HDR = "C00000"
ROJO_BG = "F8CBAD"
AZUL_HDR = "2E74B5"
AZUL_BG = "DDEBF7"
AMARILLO = "FFFF00"
BLANCO = "FFFFFF"

FONT_NAME = "Calibri"
THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

N_DIAS = 7
N_SUB = 5  # Fecha Guardia | Cumple-Si | Cumple-No | Cubre | Cubre
COL_LABEL = 1  # columna A = etiquetas (OBAC / Maquinista / VOLUNTARIOS / Oficial de)


def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def set_cell(ws, row, col, value="", bold=False, size=9, color="000000",
             bg=None, align="center", wrap=True, valign="center", rotate=0):
    c = ws.cell(row=row, column=col, value=value if value != "" else None)
    c.font = Font(name=FONT_NAME, bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap,
                             text_rotation=rotate)
    if bg:
        c.fill = fill(bg)
    c.border = BORDER_ALL
    return c


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calendario Guardia"

    # ---- Anchos de columna
    ws.column_dimensions[get_column_letter(COL_LABEL)].width = 11
    for d in range(N_DIAS):
        base = COL_LABEL + 1 + d * N_SUB
        ws.column_dimensions[get_column_letter(base)].width = 15      # Fecha
        for k in range(1, 5):
            ws.column_dimensions[get_column_letter(base + k)].width = 4.2  # checkboxes

    # ---- Título
    total_cols = COL_LABEL + N_DIAS * N_SUB
    merge(ws, 1, 1, 1, total_cols)
    set_cell(ws, 1, 1, TITULO, bold=True, size=13, bg=None)
    ws.row_dimensions[1].height = 22

    row = 3  # empieza la tabla

    # ---- Encabezado general (una sola vez), 2 filas
    r0, r1 = row, row + 1
    merge(ws, r0, COL_LABEL, r1, COL_LABEL)
    set_cell(ws, r0, COL_LABEL, "")

    for d in range(N_DIAS):
        base = COL_LABEL + 1 + d * N_SUB
        # Fecha Guardia (merge vertical)
        merge(ws, r0, base, r1, base)
        set_cell(ws, r0, base, "Fecha\nGuardia", bold=True, size=8)
        # Cumple (merge horizontal fila r0), Si/No en r1
        merge(ws, r0, base + 1, r0, base + 2)
        set_cell(ws, r0, base + 1, "Cumple", bold=True, size=8)
        set_cell(ws, r1, base + 1, "Si", bold=True, size=8)
        set_cell(ws, r1, base + 2, "No", bold=True, size=8)
        # Cubre (merge horizontal fila r0), r1 en blanco
        merge(ws, r0, base + 3, r0, base + 4)
        set_cell(ws, r0, base + 3, "Cubre", bold=True, size=8)
        set_cell(ws, r1, base + 3, "")
        set_cell(ws, r1, base + 4, "")

    ws.row_dimensions[r0].height = 15
    ws.row_dimensions[r1].height = 15
    row = r1 + 1

    # ---- Bloques semanales
    for semana in SEMANAS:
        hdr = ROJO_HDR if semana["color"] == "roja" else AZUL_HDR
        bg = ROJO_BG if semana["color"] == "roja" else AZUL_BG
        dias = semana["dias"]
        max_vol = max((len(d["voluntarios"]) for d in dias), default=0)
        max_vol = max(max_vol, 1)

        r_dia = row
        r_obac = row + 1
        r_maq = row + 2
        r_vol0 = row + 3
        r_oficial = row + 3 + max_vol

        # columna label vacía en fila de día
        set_cell(ws, r_dia, COL_LABEL, "")
        set_cell(ws, r_obac, COL_LABEL, "OBAC", bold=True, size=8, align="left")
        set_cell(ws, r_maq, COL_LABEL, "Maquinista", bold=True, size=8, align="left")

        if max_vol > 0:
            merge(ws, r_vol0, COL_LABEL, r_vol0 + max_vol - 1, COL_LABEL)
            set_cell(ws, r_vol0, COL_LABEL, "VOLUNTARIOS", bold=True, size=8, rotate=90)

        for col, dia in enumerate(dias):
            base = COL_LABEL + 1 + col * N_SUB

            # fila de día: coloreada en las 5 subcolumnas
            merge(ws, r_dia, base, r_dia, base + N_SUB - 1)
            set_cell(ws, r_dia, base, f"{dia['nombre']}\n{dia['fecha']}",
                     bold=True, size=8, color=BLANCO, bg=hdr)
            for k in range(1, N_SUB):
                set_cell(ws, r_dia, base + k, "", bg=hdr)

            # OBAC
            set_cell(ws, r_obac, base, dia["obac"], size=8, bg=bg)
            for k in range(1, N_SUB):
                set_cell(ws, r_obac, base + k, "", bg=bg)

            # Maquinista
            set_cell(ws, r_maq, base, dia["maquinista"], size=8, bg=bg)
            for k in range(1, N_SUB):
                set_cell(ws, r_maq, base + k, "", bg=bg)

            # Voluntarios
            for i in range(max_vol):
                nombre = dia["voluntarios"][i] if i < len(dia["voluntarios"]) else ""
                set_cell(ws, r_vol0 + i, base, nombre, size=8, bg=bg)
                for k in range(1, N_SUB):
                    set_cell(ws, r_vol0 + i, base + k, "", bg=bg)

        # Oficial de
        set_cell(ws, r_oficial, COL_LABEL, "Oficial de", bold=True, size=8, align="left", bg=AMARILLO)
        merge(ws, r_oficial, COL_LABEL + 1, r_oficial, total_cols)
        set_cell(ws, r_oficial, COL_LABEL + 1, semana["oficial_de"], bold=True, size=9, align="left")

        row = r_oficial + 1

    ws.freeze_panes = "B4"

    # Configuración de impresión: horizontal, ajustado a 1 hoja de ancho
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

    wb.save("calendario_guardia.xlsx")
    print("Generado: calendario_guardia.xlsx")


if __name__ == "__main__":
    build()
